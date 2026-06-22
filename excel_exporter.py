import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styles definition matching Hulu Green theme
FONT_NAME = "Arial"
COLOR_DARK_BG = "121212"      # Very dark charcoal
COLOR_NEON_GREEN = "1CE783"   # Hulu Neon Green
COLOR_FOREST_GREEN = "1C3B2B" # Forest Green
COLOR_ZEBRA_TINT = "F4FBF7"   # Very light green zebra striping
COLOR_BORDER_GREY = "D3D3D3"  # Light grey border
COLOR_BORDER_GREEN = "A2E8C4" # Light green border for summary sections
COLOR_TEXT_MUTED = "555555"

font_title = Font(name=FONT_NAME, size=16, bold=True, color=COLOR_NEON_GREEN)
font_card_title = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
font_header = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
font_body = Font(name=FONT_NAME, size=10)
font_body_bold = Font(name=FONT_NAME, size=10, bold=True)
font_body_italic = Font(name=FONT_NAME, size=9, italic=True, color=COLOR_TEXT_MUTED)

fill_dark_header = PatternFill(fill_type="solid", start_color=COLOR_DARK_BG, end_color=COLOR_DARK_BG)
fill_table_header = PatternFill(fill_type="solid", start_color=COLOR_FOREST_GREEN, end_color=COLOR_FOREST_GREEN)
fill_zebra = PatternFill(fill_type="solid", start_color=COLOR_ZEBRA_TINT, end_color=COLOR_ZEBRA_TINT)
fill_card_label = PatternFill(fill_type="solid", start_color="EAEAEA", end_color="EAEAEA")
fill_green_highlight = PatternFill(fill_type="solid", start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border_side = Side(style="thin", color=COLOR_BORDER_GREY)
border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

double_bottom_side = Side(style="double", color="000000")
thin_top_side = Side(style="thin", color="000000")
border_total = Border(top=thin_top_side, bottom=double_bottom_side)

green_border_side = Side(style="medium", color=COLOR_NEON_GREEN)
border_card = Border(left=green_border_side, right=green_border_side, top=green_border_side, bottom=green_border_side)

FORMAT_CURRENCY = "$#,##0.00"
FORMAT_PERCENT = "0.00%"
FORMAT_INT = "0"

def create_styled_workbook(pmt, annual_rate, term, calc_results):
    """
    Generate the Excel file with IFRS 16 and US GAAP schedules, formatted in Hulu Green.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 1. Summary & Comparison Sheet
    ws_summary = wb.create_sheet(title="Summary & Comparison")
    ws_summary.views.sheetView[0].showGridLines = True
    build_summary_sheet(ws_summary, pmt, annual_rate, term, calc_results)
    
    # 2. IFRS 16 Amortization Sheet
    ws_ifrs = wb.create_sheet(title="Schedule as per IFRS 16")
    ws_ifrs.views.sheetView[0].showGridLines = True
    build_ifrs_sheet(ws_ifrs, calc_results['ifrs_schedule'])
    
    # 3. US GAAP ASC 842 Amortization Sheet
    ws_us = wb.create_sheet(title="Schedule as per US GAAP")
    ws_us.views.sheetView[0].showGridLines = True
    build_us_sheet(ws_us, calc_results['us_schedule'])
    
    return wb

def build_summary_sheet(ws, pmt, annual_rate, term, calc_results):
    # Enable column auto-width logic at the end
    ws.column_dimensions['A'].width = 3
    
    # Title Block (Dark header with neon green text)
    ws.merge_cells("B2:K3")
    title_cell = ws["B2"]
    title_cell.value = " DilipFin - Lease Accounting Comparison & Adjustment Report"
    title_cell.font = font_title
    title_cell.fill = fill_dark_header
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Apply background fill to all merged title cells so borders/fill work
    for r in range(2, 4):
        for c in range(2, 12):
            ws.cell(row=r, column=c).fill = fill_dark_header
            
    # Lease parameters input card block
    ws["B5"] = "LEASE INPUTS"
    ws["B5"].font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
    
    inputs = [
        ("Monthly Rent (PMT)", pmt, FORMAT_CURRENCY),
        ("Lease Term (n months)", term, FORMAT_INT),
        ("Annual Discount Rate", annual_rate, FORMAT_PERCENT),
        ("Monthly Rate (r)", annual_rate / 12.0, FORMAT_PERCENT),
        ("Lease Classification", "Operating (US GAAP)", None)
    ]
    
    for i, (label, val, fmt) in enumerate(inputs):
        row = 6 + i
        ws.cell(row=row, column=2, value=label).font = font_body_bold
        val_cell = ws.cell(row=row, column=3, value=val)
        val_cell.font = font_body
        val_cell.alignment = align_right if fmt else align_left
        if fmt:
            val_cell.number_format = fmt
            
    # Computed PV highlight card
    ws.merge_cells("E5:F9")
    pv_card_bg = ws.cell(row=5, column=5)
    pv_card_bg.value = f"Computed PV\n\n${calc_results['pv']:,.2f}\n\nInitial Asset & Liability"
    pv_card_bg.font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_DARK_BG)
    pv_card_bg.fill = fill_green_highlight
    pv_card_bg.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for r in range(5, 10):
        for c in range(5, 7):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_green_highlight
            # Apply border around the card
            left = Side(style="medium", color=COLOR_FOREST_GREEN) if c == 5 else None
            right = Side(style="medium", color=COLOR_FOREST_GREEN) if c == 6 else None
            top = Side(style="medium", color=COLOR_FOREST_GREEN) if r == 5 else None
            bottom = Side(style="medium", color=COLOR_FOREST_GREEN) if r == 9 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Side-by-Side Yearly Comparison Section
    ws.cell(row=12, column=2, value="YEARLY EXPENSE & BALANCE COMPARISON").font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
    
    headers = [
        "Year", "Months", 
        "IFRS Expense\n(Interest + Dep)", "US GAAP Expense\n(Straight-line)", "Expense Diff\n(Debit to P&L)",
        "IFRS ROU Asset\n(Closing Balance)", "US GAAP ROU Asset\n(Closing Balance)", "ROU Asset Diff\n(Credit to ROU)",
        "IFRS Lease Liab\n(Closing Balance)", "US GAAP Lease Liab\n(Closing Balance)", "Lease Liab Diff"
    ]
    
    for c_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=13, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_table_header
        cell.alignment = align_header
        cell.border = border_data
    ws.row_dimensions[13].height = 30
    
    # Populate comparison data
    for r_idx, y_data in enumerate(calc_results['yearly_comparison'], start=14):
        # Zebra striping
        fill_row = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)
        
        row_values = [
            y_data['year'],
            y_data['months'],
            y_data['ifrs_expense'],
            y_data['us_expense'],
            y_data['expense_diff'],
            y_data['ifrs_rou'],
            y_data['us_rou'],
            y_data['rou_diff'],
            y_data['ifrs_liab'],
            y_data['us_liab'],
            y_data['liab_diff']
        ]
        
        for col_idx, val in enumerate(row_values, start=2):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.fill = fill_row
            cell.border = border_data
            
            # Format numbers
            if col_idx in [2, 3]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                cell.number_format = FORMAT_CURRENCY
                
    # Add a Totals Row
    tot_row = 14 + len(calc_results['yearly_comparison'])
    ws.cell(row=tot_row, column=2, value="Total").font = font_body_bold
    ws.cell(row=tot_row, column=2).border = border_total
    ws.cell(row=tot_row, column=3).border = border_total
    
    sum_columns = [
        (4, sum(y['ifrs_expense'] for y in calc_results['yearly_comparison'])),
        (5, sum(y['us_expense'] for y in calc_results['yearly_comparison'])),
        (6, sum(y['expense_diff'] for y in calc_results['yearly_comparison']))
    ]
    
    for c_idx, val in sum_columns:
        cell = ws.cell(row=tot_row, column=c_idx, value=val)
        cell.font = font_body_bold
        cell.number_format = FORMAT_CURRENCY
        cell.alignment = align_right
        cell.border = border_total
        
    for c_idx in range(7, 13):
        # Empty cells on total row for closing balance comparison get border
        ws.cell(row=tot_row, column=c_idx).border = border_total

    # Journal Entries section
    start_je_row = tot_row + 3
    ws.cell(row=start_je_row, column=2, value="RECOMMENDED IFRS TRANSITION JOURNAL ENTRIES").font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
    
    je_row = start_je_row + 1
    for je_group in calc_results['journal_entries']:
        ws.cell(row=je_row, column=2, value=je_group['type']).font = font_body_bold
        ws.cell(row=je_row, column=2).font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_FOREST_GREEN)
        
        # Headers for journal entry table
        ws.cell(row=je_row+1, column=2, value="Account / GL Description").font = font_header
        ws.cell(row=je_row+1, column=2).fill = fill_table_header
        ws.cell(row=je_row+1, column=2).border = border_data
        
        ws.cell(row=je_row+1, column=3, value="Debit").font = font_header
        ws.cell(row=je_row+1, column=3).fill = fill_table_header
        ws.cell(row=je_row+1, column=3).border = border_data
        ws.cell(row=je_row+1, column=3).alignment = align_center
        
        ws.cell(row=je_row+1, column=4, value="Credit").font = font_header
        ws.cell(row=je_row+1, column=4).fill = fill_table_header
        ws.cell(row=je_row+1, column=4).border = border_data
        ws.cell(row=je_row+1, column=4).alignment = align_center
        
        item_row = je_row + 2
        for entry in je_group['entries']:
            # Indent credit accounts slightly
            acct = entry['account']
            if entry['credit'] > 0:
                acct = "    " + acct
            ws.cell(row=item_row, column=2, value=acct).font = font_body
            ws.cell(row=item_row, column=2).border = border_data
            
            deb_cell = ws.cell(row=item_row, column=3, value=entry['debit'] if entry['debit'] > 0 else "")
            deb_cell.font = font_body
            deb_cell.number_format = FORMAT_CURRENCY
            deb_cell.alignment = align_right
            deb_cell.border = border_data
            
            cred_cell = ws.cell(row=item_row, column=4, value=entry['credit'] if entry['credit'] > 0 else "")
            cred_cell.font = font_body
            cred_cell.number_format = FORMAT_CURRENCY
            cred_cell.alignment = align_right
            cred_cell.border = border_data
            
            item_row += 1
            
        je_row = item_row + 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        # Ignore merged title row and parameters card for width calculation
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'A':
            continue
        for cell in col:
            if cell.row > 3 and cell.value and not isinstance(cell.value, str) and not str(cell.value).startswith('='):
                val_str = f"{cell.value:,.2f}" if isinstance(cell.value, (int, float)) else str(cell.value)
                max_len = max(max_len, len(val_str))
            elif cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Extra spacing for specific description columns
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

def build_ifrs_sheet(ws, schedule):
    # Header
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = " Lease Amortization Schedule - IFRS 16"
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_NEON_GREEN)
    title_cell.fill = fill_dark_header
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = fill_dark_header
    ws.row_dimensions[1].height = 25
    
    headers = [
        "Month", "Cash Payment", "Interest Expense", "Depreciation Expense", 
        "Total P&L Expense", "Lease Liability (Closing)", "ROU Asset (Closing)"
    ]
    
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_table_header
        cell.alignment = align_header
        cell.border = border_data
    ws.row_dimensions[3].height = 25
    
    for r_idx, row_data in enumerate(schedule, start=4):
        fill_row = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)
        
        row_values = [
            row_data['month'],
            row_data['cash'],
            row_data['interest'],
            row_data['depreciation'],
            row_data['total_expense'],
            row_data['liability_closing'],
            row_data['rou_closing']
        ]
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.fill = fill_row
            cell.border = border_data
            
            if col_idx == 1:
                cell.alignment = align_center
                cell.number_format = FORMAT_INT
            else:
                cell.alignment = align_right
                if row_data['month'] == 0 and col_idx in [2, 3, 4, 5]:
                    # For Month 0, Cash, Interest, Dep, and Total exp are empty/hyphen
                    cell.value = "-"
                    cell.alignment = align_center
                else:
                    cell.number_format = FORMAT_CURRENCY
                    
    # Totals Row
    tot_row = len(schedule) + 4
    ws.cell(row=tot_row, column=1, value="Total").font = font_body_bold
    ws.cell(row=tot_row, column=1).border = border_total
    
    sum_columns = [
        (2, sum(item['cash'] for item in schedule)),
        (3, sum(item['interest'] for item in schedule)),
        (4, sum(item['depreciation'] for item in schedule)),
        (5, sum(item['total_expense'] for item in schedule))
    ]
    
    for c_idx, val in sum_columns:
        cell = ws.cell(row=tot_row, column=c_idx, value=val)
        cell.font = font_body_bold
        cell.number_format = FORMAT_CURRENCY
        cell.alignment = align_right
        cell.border = border_total
        
    for c_idx in [6, 7]:
        ws.cell(row=tot_row, column=c_idx).border = border_total
        
    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1 and cell.value and not str(cell.value).startswith('='):
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

def build_us_sheet(ws, schedule):
    # Header
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = " Lease Amortization Schedule - US GAAP ASC 842"
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_NEON_GREEN)
    title_cell.fill = fill_dark_header
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 9):
        ws.cell(row=1, column=c).fill = fill_dark_header
    ws.row_dimensions[1].height = 25
    
    headers = [
        "Month", "Cash Payment", "Lease Expense", "Interest Expense (A)", 
        "Liability Reduction (B)", "ROU Asset Amortization (C)", 
        "Liability Balance", "ROU Asset Balance"
    ]
    
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_table_header
        cell.alignment = align_header
        cell.border = border_data
    ws.row_dimensions[3].height = 25
    
    for r_idx, row_data in enumerate(schedule, start=4):
        fill_row = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)
        
        row_values = [
            row_data['month'],
            row_data['cash'],
            row_data['lease_expense'],
            row_data['interest'],
            row_data['liab_reduction'],
            row_data['rou_amortization'],
            row_data['liability_balance'],
            row_data['rou_balance']
        ]
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.fill = fill_row
            cell.border = border_data
            
            if col_idx == 1:
                cell.alignment = align_center
                cell.number_format = FORMAT_INT
            else:
                cell.alignment = align_right
                if row_data['month'] == 0 and col_idx in [2, 3, 4, 5, 6]:
                    cell.value = "-"
                    cell.alignment = align_center
                else:
                    cell.number_format = FORMAT_CURRENCY
                    
    # Totals Row
    tot_row = len(schedule) + 4
    ws.cell(row=tot_row, column=1, value="Total").font = font_body_bold
    ws.cell(row=tot_row, column=1).border = border_total
    
    sum_columns = [
        (2, sum(item['cash'] for item in schedule)),
        (3, sum(item['lease_expense'] for item in schedule)),
        (4, sum(item['interest'] for item in schedule)),
        (5, sum(item['liab_reduction'] for item in schedule)),
        (6, sum(item['rou_amortization'] for item in schedule))
    ]
    
    for c_idx, val in sum_columns:
        cell = ws.cell(row=tot_row, column=c_idx, value=val)
        cell.font = font_body_bold
        cell.number_format = FORMAT_CURRENCY
        cell.alignment = align_right
        cell.border = border_total
        
    for c_idx in [7, 8]:
        ws.cell(row=tot_row, column=c_idx).border = border_total
        
    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1 and cell.value and not str(cell.value).startswith('='):
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

def create_portfolio_workbook(portfolio_results):
    """
    Generate an Excel workbook for a list of multiple leases consolidated.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 1. Portfolio Summary Sheet
    ws_summary = wb.create_sheet(title="Portfolio Summary")
    ws_summary.views.sheetView[0].showGridLines = True
    build_portfolio_summary_sheet(ws_summary, portfolio_results)
    
    # 2. Consolidated IFRS 16 Schedules
    ws_ifrs = wb.create_sheet(title="IFRS 16 Schedules")
    ws_ifrs.views.sheetView[0].showGridLines = True
    
    ws_ifrs.merge_cells("A1:G1")
    title_cell = ws_ifrs["A1"]
    title_cell.value = " Lease Amortization Schedules - IFRS 16 (Portfolio)"
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_NEON_GREEN)
    title_cell.fill = fill_dark_header
    title_cell.alignment = align_left
    for c in range(1, 8):
        ws_ifrs.cell(row=1, column=c).fill = fill_dark_header
    ws_ifrs.row_dimensions[1].height = 25
    
    # 3. Consolidated US GAAP Schedules
    ws_us = wb.create_sheet(title="US GAAP Schedules")
    ws_us.views.sheetView[0].showGridLines = True
    
    ws_us.merge_cells("A1:H1")
    title_cell_us = ws_us["A1"]
    title_cell_us.value = " Lease Amortization Schedules - US GAAP ASC 842 (Portfolio)"
    title_cell_us.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_NEON_GREEN)
    title_cell_us.fill = fill_dark_header
    title_cell_us.alignment = align_left
    for c in range(1, 9):
        ws_us.cell(row=1, column=c).fill = fill_dark_header
    ws_us.row_dimensions[1].height = 25

    # Stack individual schedules
    r_idx_ifrs = 3
    r_idx_us = 3
    
    headers_ifrs = [
        "Month", "Cash Payment", "Interest Expense", "Depreciation Expense", 
        "Total P&L Expense", "Lease Liability (Closing)", "ROU Asset (Closing)"
    ]
    
    headers_us = [
        "Month", "Cash Payment", "Lease Expense", "Interest Expense (A)", 
        "Liability Reduction (B)", "ROU Asset Amortization (C)", 
        "Liability Balance", "ROU Asset Balance"
    ]
    
    for name, res in portfolio_results['individual_results'].items():
        # --- IFRS SHEET STACK ---
        ws_ifrs.cell(row=r_idx_ifrs, column=1, value=f"Lease: {name}").font = font_body_bold
        ws_ifrs.cell(row=r_idx_ifrs, column=1).font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
        r_idx_ifrs += 1
        
        for c_idx, h in enumerate(headers_ifrs, start=1):
            cell = ws_ifrs.cell(row=r_idx_ifrs, column=c_idx, value=h)
            cell.font = font_header
            cell.fill = fill_table_header
            cell.alignment = align_header
            cell.border = border_data
        ws_ifrs.row_dimensions[r_idx_ifrs].height = 22
        r_idx_ifrs += 1
        
        for m_data in res['ifrs_schedule']:
            fill_row = fill_zebra if r_idx_ifrs % 2 == 1 else PatternFill(fill_type=None)
            row_values = [
                m_data['month'], m_data['cash'], m_data['interest'], m_data['depreciation'],
                m_data['total_expense'], m_data['liability_closing'], m_data['rou_closing']
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws_ifrs.cell(row=r_idx_ifrs, column=col_idx, value=val)
                cell.font = font_body
                cell.fill = fill_row
                cell.border = border_data
                if col_idx == 1:
                    cell.alignment = align_center
                    cell.number_format = FORMAT_INT
                else:
                    cell.alignment = align_right
                    if m_data['month'] == 0 and col_idx in [2, 3, 4, 5]:
                        cell.value = "-"
                        cell.alignment = align_center
                    else:
                        cell.number_format = FORMAT_CURRENCY
            r_idx_ifrs += 1
            
        r_idx_ifrs += 2 # gap
        
        # --- US GAAP SHEET STACK ---
        ws_us.cell(row=r_idx_us, column=1, value=f"Lease: {name}").font = font_body_bold
        ws_us.cell(row=r_idx_us, column=1).font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
        r_idx_us += 1
        
        for c_idx, h in enumerate(headers_us, start=1):
            cell = ws_us.cell(row=r_idx_us, column=c_idx, value=h)
            cell.font = font_header
            cell.fill = fill_table_header
            cell.alignment = align_header
            cell.border = border_data
        ws_us.row_dimensions[r_idx_us].height = 22
        r_idx_us += 1
        
        for m_data in res['us_schedule']:
            fill_row = fill_zebra if r_idx_us % 2 == 1 else PatternFill(fill_type=None)
            row_values = [
                m_data['month'], m_data['cash'], m_data['lease_expense'], m_data['interest'],
                m_data['liab_reduction'], m_data['rou_amortization'], m_data['liability_balance'],
                m_data['rou_balance']
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws_us.cell(row=r_idx_us, column=col_idx, value=val)
                cell.font = font_body
                cell.fill = fill_row
                cell.border = border_data
                if col_idx == 1:
                    cell.alignment = align_center
                    cell.number_format = FORMAT_INT
                else:
                    cell.alignment = align_right
                    if m_data['month'] == 0 and col_idx in [2, 3, 4, 5, 6]:
                        cell.value = "-"
                        cell.alignment = align_center
                    else:
                        cell.number_format = FORMAT_CURRENCY
            r_idx_us += 1
            
        r_idx_us += 2 # gap
        
    # Auto-adjust column widths for schedules
    for ws in [ws_ifrs, ws_us]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 1 and cell.value and not str(cell.value).startswith('='):
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
            
    return wb

def build_portfolio_summary_sheet(ws, results):
    # Title Block
    ws.merge_cells("B2:K3")
    title_cell = ws["B2"]
    title_cell.value = " DilipFin - Portfolio Lease Comparison & Transition Report"
    title_cell.font = font_title
    title_cell.fill = fill_dark_header
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for r in range(2, 4):
        for c in range(2, 12):
            ws.cell(row=r, column=c).fill = fill_dark_header
            
    # List of leases included in report
    ws.cell(row=5, column=2, value="INCLUDED LEASES").font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
    
    lease_headers = ["Lease Name", "Monthly Rent (PMT)", "Lease Term (n)", "Annual Discount Rate"]
    for c_idx, h in enumerate(lease_headers, start=2):
        cell = ws.cell(row=6, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_table_header
        cell.alignment = align_header
        cell.border = border_data
    ws.row_dimensions[6].height = 22
    
    for idx, lease in enumerate(results['leases']):
        r_idx = 7 + idx
        fill_row = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)
        
        ws.cell(row=r_idx, column=2, value=lease.get('name', f"Lease {idx+1}")).font = font_body
        ws.cell(row=r_idx, column=2).border = border_data
        ws.cell(row=r_idx, column=2).fill = fill_row
        
        c2 = ws.cell(row=r_idx, column=3, value=lease['rent'])
        c2.font = font_body
        c2.number_format = FORMAT_CURRENCY
        c2.alignment = align_right
        c2.border = border_data
        c2.fill = fill_row
        
        c3 = ws.cell(row=r_idx, column=4, value=lease['term'])
        c3.font = font_body
        c3.number_format = FORMAT_INT
        c3.alignment = align_center
        c3.border = border_data
        c3.fill = fill_row
        
        c4 = ws.cell(row=r_idx, column=5, value=lease['rate'])
        c4.font = font_body
        c4.number_format = FORMAT_PERCENT
        c4.alignment = align_right
        c4.border = border_data
        c4.fill = fill_row
        
    # Portfolio Totals card
    tot_card_row = 7 + len(results['leases']) + 2
    ws.cell(row=tot_card_row, column=2, value="PORTFOLIO TOTALS").font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
    
    ws.cell(row=tot_card_row+1, column=2, value="Combined Present Value (PV)").font = font_body_bold
    ws.cell(row=tot_card_row+1, column=3, value=results['pv']).font = font_body_bold
    ws.cell(row=tot_card_row+1, column=3).number_format = FORMAT_CURRENCY
    ws.cell(row=tot_card_row+1, column=3).alignment = align_right
    ws.cell(row=tot_card_row+1, column=2).border = border_total
    ws.cell(row=tot_card_row+1, column=3).border = border_total
    
    # Side-by-Side Yearly Comparison Section
    y_start_row = tot_card_row + 4
    ws.cell(row=y_start_row, column=2, value="PORTFOLIO CONSOLIDATED YEARLY COMPARISON").font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
    
    headers = [
        "Year", "Months", 
        "IFRS Expense\n(Interest + Dep)", "US GAAP Expense\n(Straight-line)", "Expense Diff\n(Debit to P&L)",
        "IFRS ROU Asset\n(Closing Balance)", "US GAAP ROU Asset\n(Closing Balance)", "ROU Asset Diff\n(Credit to ROU)",
        "IFRS Lease Liab\n(Closing Balance)", "US GAAP Lease Liab\n(Closing Balance)", "Lease Liab Diff"
    ]
    
    for c_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=y_start_row+1, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_table_header
        cell.alignment = align_header
        cell.border = border_data
    ws.row_dimensions[y_start_row+1].height = 30
    
    # Populate comparison data
    for r_offset, y_data in enumerate(results['yearly_comparison']):
        r_idx = y_start_row + 2 + r_offset
        fill_row = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)
        
        row_values = [
            y_data['year'], y_data['months'], y_data['ifrs_expense'], y_data['us_expense'],
            y_data['expense_diff'], y_data['ifrs_rou'], y_data['us_rou'], y_data['rou_diff'],
            y_data['ifrs_liab'], y_data['us_liab'], y_data['liab_diff']
        ]
        
        for col_idx, val in enumerate(row_values, start=2):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.fill = fill_row
            cell.border = border_data
            if col_idx in [2, 3]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                cell.number_format = FORMAT_CURRENCY
                
    # Add a Totals Row
    tot_row = y_start_row + 2 + len(results['yearly_comparison'])
    ws.cell(row=tot_row, column=2, value="Total").font = font_body_bold
    ws.cell(row=tot_row, column=2).border = border_total
    ws.cell(row=tot_row, column=3).border = border_total
    
    sum_columns = [
        (4, sum(y['ifrs_expense'] for y in results['yearly_comparison'])),
        (5, sum(y['us_expense'] for y in results['yearly_comparison'])),
        (6, sum(y['expense_diff'] for y in results['yearly_comparison']))
    ]
    
    for c_idx, val in sum_columns:
        cell = ws.cell(row=tot_row, column=c_idx, value=val)
        cell.font = font_body_bold
        cell.number_format = FORMAT_CURRENCY
        cell.alignment = align_right
        cell.border = border_total
        
    for c_idx in range(7, 13):
        ws.cell(row=tot_row, column=c_idx).border = border_total

    # Journal Entries section
    start_je_row = tot_row + 3
    ws.cell(row=start_je_row, column=2, value="RECOMMENDED PORTFOLIO IFRS TRANSITION JOURNAL ENTRIES").font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_FOREST_GREEN)
    
    je_row = start_je_row + 1
    for je_group in results['journal_entries']:
        ws.cell(row=je_row, column=2, value=je_group['type']).font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_FOREST_GREEN)
        
        ws.cell(row=je_row+1, column=2, value="Account / GL Description").font = font_header
        ws.cell(row=je_row+1, column=2).fill = fill_table_header
        ws.cell(row=je_row+1, column=2).border = border_data
        
        ws.cell(row=je_row+1, column=3, value="Debit").font = font_header
        ws.cell(row=je_row+1, column=3).fill = fill_table_header
        ws.cell(row=je_row+1, column=3).border = border_data
        ws.cell(row=je_row+1, column=3).alignment = align_center
        
        ws.cell(row=je_row+1, column=4, value="Credit").font = font_header
        ws.cell(row=je_row+1, column=4).fill = fill_table_header
        ws.cell(row=je_row+1, column=4).border = border_data
        ws.cell(row=je_row+1, column=4).alignment = align_center
        
        item_row = je_row + 2
        for entry in je_group['entries']:
            acct = entry['account']
            if entry['credit'] > 0:
                acct = "    " + acct
            ws.cell(row=item_row, column=2, value=acct).font = font_body
            ws.cell(row=item_row, column=2).border = border_data
            
            deb_cell = ws.cell(row=item_row, column=3, value=entry['debit'] if entry['debit'] > 0 else "")
            deb_cell.font = font_body
            deb_cell.number_format = FORMAT_CURRENCY
            deb_cell.alignment = align_right
            deb_cell.border = border_data
            
            cred_cell = ws.cell(row=item_row, column=4, value=entry['credit'] if entry['credit'] > 0 else "")
            cred_cell.font = font_body
            cred_cell.number_format = FORMAT_CURRENCY
            cred_cell.alignment = align_right
            cred_cell.border = border_data
            
            item_row += 1
            
        je_row = item_row + 1
        
    # Set custom widths for key summary columns
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
