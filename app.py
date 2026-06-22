import os
import io
from flask import Flask, request, jsonify, render_template, send_file
import openpyxl
from financial_engine import perform_comparison, perform_portfolio_comparison
from excel_exporter import create_styled_workbook, create_portfolio_workbook

app = Flask(__name__, template_folder='templates', static_folder='static')

# Secret key for sessions
app.secret_key = 'dilipfin_secret_key'

def parse_excel_file(filepath_or_stream):
    """
    Parse Excel file to find lease parameters.
    Supports both:
    1. Table format (columns for Lease, Rent, Rate, Tenure).
    2. Key-Value format (isolated cells like 'Rent: 10000', 'Tenure: 24', etc.).
    """
    wb = openpyxl.load_workbook(filepath_or_stream, data_only=True)
    leases = []
    
    # Let's inspect each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # 1. Search for Table Format
        # We look for a header row containing column headers like Rent/PMT, tenure/term, rate/interest.
        table_headers = None
        header_row_idx = -1
        col_mappings = {} # maps 'rent', 'rate', 'term', 'name' to column index
        
        for r_idx in range(1, min(sheet.max_row + 1, 20)):
            row_vals = [str(sheet.cell(row=r_idx, column=c_idx).value).strip().lower() 
                        for c_idx in range(1, min(sheet.max_column + 1, 15))]
            
            # Check if this row looks like a lease table header
            has_rent = any('rent' in val or 'pmt' in val for val in row_vals if val)
            has_rate = any('rate' in val or 'interest' in val or 'discount' in val for val in row_vals if val)
            has_term = any('tenure' in val or 'term' in val or 'period' in val or 'months' in val for val in row_vals if val)
            
            if has_rent and (has_rate or has_term):
                table_headers = row_vals
                header_row_idx = r_idx
                
                # Determine columns
                for c_idx, val in enumerate(row_vals):
                    if not val:
                        continue
                    if 'rent' in val or 'pmt' in val:
                        col_mappings['rent'] = c_idx + 1
                    elif 'rate' in val or 'interest' in val or 'discount' in val:
                        col_mappings['rate'] = c_idx + 1
                    elif 'tenure' in val or 'term' in val or 'period' in val or 'months' in val:
                        col_mappings['term'] = c_idx + 1
                    elif 'lease' in val or 'name' in val or 'list' in val:
                        col_mappings['name'] = c_idx + 1
                break
                
        if table_headers and 'rent' in col_mappings:
            # Parse rows below the header
            for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
                name_val = sheet.cell(row=r_idx, column=col_mappings.get('name', 1)).value
                rent_val = sheet.cell(row=r_idx, column=col_mappings['rent']).value
                rate_val = sheet.cell(row=r_idx, column=col_mappings.get('rate', 1)).value
                term_val = sheet.cell(row=r_idx, column=col_mappings.get('term', 1)).value
                
                # If rent is empty, we stop parsing this table
                if rent_val is None or str(rent_val).strip() == "":
                    break
                    
                # Clean and convert values
                try:
                    rent = float(str(rent_val).replace('$', '').replace(',', '').strip())
                    term = int(float(str(term_val).strip())) if term_val is not None else 12
                    
                    if rate_val is not None:
                        rate_str = str(rate_val).replace('%', '').strip()
                        rate = float(rate_str)
                        # If interest is > 1 (e.g. 5 instead of 0.05), divide by 100
                        if rate >= 1.0 or ('%' in str(rate_val)):
                            # Check if it was formatted as percentage in Excel
                            # sometimes cell.value is float (0.05) and rate_str is '0.05', so check raw rate_val
                            if isinstance(rate_val, float) and rate_val < 1.0:
                                rate = rate_val
                            else:
                                if rate >= 1.0:
                                    rate = rate / 100.0
                    else:
                        rate = 0.0
                        
                    lease_name = str(name_val).strip() if name_val is not None else f"Lease {len(leases) + 1}"
                    
                    leases.append({
                        'name': f"{sheet_name} - {lease_name}",
                        'rent': rent,
                        'rate': rate,
                        'term': term
                    })
                except Exception as e:
                    # Skip rows that fail to parse
                    continue
            # If we found leases, continue to next sheet
            if leases:
                continue
                
        # 2. Search for Key-Value format (isolated cells)
        # Scan cells for Rent, Term, and Interest Rate
        rent_cell = None
        term_cell = None
        rate_cell = None
        
        for r_idx in range(1, min(sheet.max_row + 1, 30)):
            for c_idx in range(1, min(sheet.max_column + 1, 10)):
                cell_val = str(sheet.cell(row=r_idx, column=c_idx).value or "").strip().lower()
                if not cell_val:
                    continue
                
                # Look for Rent
                if cell_val in ['rent', 'monthly rent', 'pmt', 'rent (pmt)', 'monthly rent (pmt)']:
                    # check adjacent cell
                    val = sheet.cell(row=r_idx, column=c_idx + 1).value
                    if val is not None:
                        rent_cell = val
                # Look for Term
                elif cell_val in ['term', 'lease term', 'tenure', 'term (n)', 'lease term (n)', 'period']:
                    val = sheet.cell(row=r_idx, column=c_idx + 1).value
                    if val is not None:
                        term_cell = val
                # Look for Interest Rate
                elif cell_val in ['rate', 'interest rate', 'annual rate', 'discount rate', 'annual discount rate', 'monthly rate (r)']:
                    # check adjacent cell
                    val = sheet.cell(row=r_idx, column=c_idx + 1).value
                    # If this is "monthly rate", it might be a formula. Let's make sure we find the annual rate.
                    if val is not None:
                        if 'monthly' in cell_val:
                            # If it's monthly, we will multiply by 12 later if we don't find annual
                            rate_cell = ('monthly', val)
                        else:
                            rate_cell = ('annual', val)
                            
        # If we found at least rent and term
        if rent_cell is not None and term_cell is not None:
            try:
                rent = float(str(rent_cell).replace('$', '').replace(',', '').strip())
                term = int(float(str(term_cell).replace('months', '').strip()))
                
                rate = 0.05 # default
                if rate_cell is not None:
                    rate_type, r_val = rate_cell
                    rate_num = float(str(r_val).replace('%', '').strip())
                    if '%' in str(r_val) or rate_num >= 1.0:
                        if isinstance(r_val, float) and r_val < 1.0:
                            rate_num = r_val
                        else:
                            if rate_num >= 1.0:
                                rate_num = rate_num / 100.0
                    if rate_type == 'monthly':
                        rate = rate_num * 12.0
                    else:
                        rate = rate_num
                        
                leases.append({
                    'name': f"{sheet_name} - Single Lease",
                    'rent': rent,
                    'rate': rate,
                    'term': term
                })
            except:
                pass

    # If no leases found, check if we can fall back to the first sheet or return empty
    return leases

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/calculate', methods=['POST'])
def calculate():
    data = request.json
    try:
        pmt = float(data.get('pmt', 10000))
        annual_rate = float(data.get('annual_rate', 0.05))
        term = int(data.get('term', 24))
        
        results = perform_comparison(pmt, annual_rate, term)
        return jsonify({
            'success': True,
            'results': results
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/parse', methods=['POST'])
def parse_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
        
    try:
        # Save file to a temporary stream and parse
        file_stream = io.BytesIO(file.read())
        leases = parse_excel_file(file_stream)
        
        if not leases:
            return jsonify({
                'success': False,
                'error': 'Could not parse any leases from the file. Please ensure it contains lease parameters (Rent, Term, Discount Rate).'
            }), 400
            
        return jsonify({
            'success': True,
            'leases': leases
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
def export_excel():
    data = request.json
    try:
        pmt = float(data.get('pmt', 10000))
        annual_rate = float(data.get('annual_rate', 0.05))
        term = int(data.get('term', 24))
        
        results = perform_comparison(pmt, annual_rate, term)
        wb = create_styled_workbook(pmt, annual_rate, term, results)
        
        # Save workbook to BytesIO stream
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='DilipFin_Lease_GAAP_IFRS_Report.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/calculate_portfolio', methods=['POST'])
def calculate_portfolio():
    data = request.json
    try:
        leases = data.get('leases', [])
        if not leases:
            return jsonify({'success': False, 'error': 'No leases provided'}), 400
        results = perform_portfolio_comparison(leases)
        return jsonify({
            'success': True,
            'results': results
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/export_portfolio', methods=['POST'])
def export_portfolio():
    data = request.json
    try:
        leases = data.get('leases', [])
        if not leases:
            return jsonify({'success': False, 'error': 'No leases provided'}), 400
        results = perform_portfolio_comparison(leases)
        wb = create_portfolio_workbook(results)
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='DilipFin_Portfolio_Lease_Report.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    # Run on port 5000
    app.run(host='127.0.0.1', port=5000, debug=True)
